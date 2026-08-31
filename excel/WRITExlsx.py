import openpyxl as xlsx
from datetime import datetime as dt, timedelta as td
import sys
sys.path.append("../HJ-API")
from general.xlsxRows import rows
from general.functions import week_range, sendEmail as email, excelYearlyWipe as resetXlsx
import excel.GETquantities as quant, excel.WRITEstock as stock

def update():
    jlpQuant, bqQuant, jlpTotal, bqTotal,totalUnits = quant.pullSF()
    workbook = xlsx.load_workbook("excel/SalesReport_TEST.xlsx")
    sheet = workbook.active
    dateRange = week_range()

    today = dt.now().date()
    jan_1 = dt(today.year, 1, 1).date()
    week_1_end = jan_1 + td(days=7)
    
    if jan_1 <= today <= week_1_end and today.weekday() == 0:  # Monday of week 1
        resetXlsx()
        workbook = xlsx.load_workbook("excel/SalesReport.xlsx")
        sheet = workbook.active


    for i,cell in enumerate(sheet[20]):
        if cell.value == dateRange:
            dateCol = sheet[20][i-1].column_letter
            break

    #jlp
    for product in jlpQuant:
        if product in rows:
            sheet[f"{dateCol}{rows[product][0]}"] = jlpQuant[product]

    #b&q
    for product in bqQuant:
        if product in rows:
            sheet[f"{dateCol}{rows[product][1]}"] = bqQuant[product]

     
    monthTotal = 0
    for col in sheet.iter_cols(min_row=59, max_row=59, max_col=52, min_col=2):
        if col[0].value != "" and col[0].value != 0:
            monthTotal = col[0].value
        elif col[0].value == 0:
            break


    workbook.save("excel/SalesReport_TEST.xlsx")

    totalQuant = jlpQuant.copy()
    total = jlpTotal + bqTotal
    totals = [total,jlpTotal,bqTotal]
    for key, value in bqQuant.items():
        totalQuant[key] = totalQuant.get(key, 0) + value
    stock.reorder(totalQuant, totals, totalUnits, monthTotal)

    # email("Sales report", "See the attached sales report", "rebecca@haywardjardine.co.uk",True)
    # email("Sales report", "See the attached sales report", "alistair@haywardjardine.co.uk",True)

update()