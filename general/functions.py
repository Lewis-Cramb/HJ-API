from datetime import date as date, timedelta as td, datetime as dt
import emails, requests as rqs
from general.productNames import miraklToSF, vsToSF, sfToXero
from general.surchargePostcodes import codes as surCodes
from base64 import b64encode as b64
import openpyxl as xlsx

def oldHeader(filename): #Use this function if you do not need the "bearer" in the auth key (So older APIs without OAuth 2.0)
    with open(f"txts/{filename}.txt") as rf:
        token = rf.read() #Tokens for sales platforms
    return {"Authorization":f"{token}"}

def headerGeneration(filename): #This function generates the headers needed for the JWT (JSON Web Token)
    with open(f"{filename}.txt") as rf:
        token = rf.read() #Tokens for sales platforms
    return {"Authorization":f"Bearer {token}"}

def vsHeader(): #this function generates the header for virtualstock
    with open("txts/vsUsername.txt", "r") as rf:
        username = rf.read().strip()
    with open("txts/vsPassword.txt", "r") as rf:
        password = rf.read().strip()

    credentials = b64(f"{username}:{password}".encode()).decode()
    return {"Authorization": f"Basic {credentials}"}
      

def startDate(): #This function is used for filtering data to be relevant and un-entered on SF
    if date.today().strftime("%A") == "Monday":
        return date.today()-td(days=3)
    return date.today()-td(days=1)


def format_date(raw_date): #this function is used to convert the date into yyyy-mm-dd
    try:
        return dt.strptime(raw_date, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return dt.strptime(raw_date, "%d-%m-%Y").strftime("%Y-%m-%d")

def formatDateOpposite(date): #this formats dd-mm-yyyy hence opposite
    comps = date.split("-")
    return f"{comps[2]}/{comps[1]}/{comps[0]}"
    
def xeroDue():
    today = date.today()
    due = today + td(days=60)
    return due.strftime("%Y-%m-%d")
    


def printOrders(orders): #This function is used to print the orders just to debug them
    for order in orders:
        print(f"On {order["orderDate"]} and from {order["accName"]}, {order["custName"]} ({order["custPO"]}) ordered:")
        for product in order["products"]:
            print(f"{order["products"][product]}x {product}")
        print(f"For a total of {order["cost"]}")
        print(f"This is to be delivered by {order["shipName"]} by {order["expDate"]}")
        print("\n")


def convertNames(data, source): #this function is used to change the names of products
    if "B&Q" in source:
        conversion = miraklToSF
    elif source == "JLP":
        conversion = vsToSF
    elif source == "SF":
        conversion = sfToXero
    for order in data:
        new_prods = {}
        for product in order["products"]:
            if "2x " in product:
                order["products"][product] = order["products"][product]*2
            elif "Pack of 4" in product:
                order["products"][product] = order["products"][product]*4
        order["products"] = {
            conversion.get(cleaned_name, cleaned_name): qty
            for name, qty in order["products"].items()
            for cleaned_name in [name if "#" not in name else name.replace("#", "")]
        }
    
    return data

def removeTitles(data):
    for order in data:
        if "Mr " in order["custName"]:
            order["custName"] = order["custName"].partition("Mr ")[2]
        elif "Mrs " in order["custName"]:
            order["custName"] = order["custName"].partition("Mrs ")[2]
        elif "Ms " in order["custName"]:
            order["custName"] = order["custName"].partition("Ms ")[2]
        elif "Dr " in order["custName"]:
            order["custName"] = order["custName"].partition("Dr ")[2]
    return data

def sendEmail(title, body, to="help@haywardjardine.co.uk", report=False):

    fullBody = f"Hi, \n {body} \n \n LewisBot \n\n (You can reply to this email, it is my personal and will always work)"
    fullTitle = f"LewisBot HJ API - {title}"

    password = open("txts/emailPassword.txt","r").read().strip()

    message = emails.html(
        text=fullBody, 
        subject=fullTitle, 
        mail_from=("Lewis", "lewiscramb@icloud.com")
    )

    if report:
        message.attach(
            filename="SalesReport.xlsx",
            data=open("excel/SalesReport.xlsx","rb")
        )
        message.attach(
            filename="StockLevels.xlsx",
            data=open("excel/StockLevels.xlsx","rb")
        )
    response = message.send(
        to=to,
        smtp={"host": "smtp.mail.icloud.com","port": 587,"tls": True,"user": "lewiscramb@icloud.com","password": password})
    pass


def shippingPostcodes(location):
    post_code = location["post_code"]
    area_code = ""
    for letter in post_code:
        if letter.isalpha():
            area_code += letter
        else:
            break

    return area_code in surCodes

def date_format(date):
    day = date.day
    month = date.strftime("%B")
    
    suffix = get_suffix(day)
    
    return f"{day}{suffix} {month}"

def get_suffix(day):
    if day in [1, 21, 31]:
        return "st"
    elif day in [2, 22]:
        return "nd"
    elif day in [3, 23]:
        return "rd"
    else:
        return "th"

def week_range():
    today = dt.now().date()

    start_date = today - td(days=today.weekday())
    end_date = start_date + td(days=6)
    
    start_formatted = date_format(start_date)
    end_formatted = date_format(end_date)
    
    if start_date.month == end_date.month:
        return f"{start_date.day}{get_suffix(start_date.day)}-{end_date.day}{get_suffix(end_date.day)} {start_date.strftime('%b')}"
    else:
        return f"{start_formatted} - {end_formatted}"

def xeroToken():
    clientID = open("txts/xeroID.txt","r").read().strip()
    clientSec = open("txts/xeroSec.txt","r").read().strip()

    header = {"Authorization" : "Basic " + b64(f"{clientID}:{clientSec}".encode()).decode()}
    info = {"grant_type" : "client_credentials", "scope":"accounting.invoices accounting.settings"}

    token = rqs.post("https://identity.xero.com/connect/token",headers=header,data=info)    
    return token.json()["access_token"]

def xeroDate():
    now = (startDate()-td(days=3)).strftime("%Y-%m-%d")
    comps = now.split("-")
    return f"DateTime({comps[0]}, {comps[1]}, {comps[2]})"


def findMax(list):
    max = list[0]
    for i in range(1, len(list)):
        if list[i] > max:
            max = list[i]
    return int(max)

def monthToCol(monthNum, year):
    conversions = {
        "1":"W", "2":"X", "3":"Y", "4":"Z", "5":"AA", "6":"AB",
        "7":"AC", "8":"AD", "9":"AE", "10":"AF", "11":"AG", "12":"AH"
    }
    if year=="this:":
        return f"{conversions[monthNum]}6"
    else:
        return f"{conversions[monthNum]}9"
    

def excelYearlyWipe():
    workbook = xlsx.load_workbook("excel/SalesReport.xlsx")
    sheet = workbook.active
    
    next_year = dt.now().year + 1

    data_ranges = [
        (21, 35),  
        (40, 54),  
    ]
    
    for start_row, end_row in data_ranges:
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                cell.value = None
    
    header_row_pairs = [
        (19, 20),
        (38, 39),
        (62, 63),
        (81, 82)
    ]
    
    jan_1 = dt(next_year, 1, 1)
    first_monday = jan_1 + td(days=(7 - jan_1.weekday()) % 7)
    
    for header_row_1, header_row_2 in header_row_pairs:
        col = 2 
        curDate = first_monday
        
        while curDate.year == next_year:
            monthHeader = curDate.strftime("%d %b")
            sheet.cell(row=header_row_1, column=col, value=monthHeader)
            
            weekEnd = curDate + td(days=6)
            weekHeader = f"{curDate.strftime('%d %b')} - {weekEnd.strftime('%d %b')}"
            sheet.cell(row=header_row_2, column=col, value=weekHeader)
            
            col += 1
            curDate += td(days=7)
    
    workbook.save("excel/SalesReport.xlsx")