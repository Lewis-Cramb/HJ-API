import openpyxl as xlsx
from datetime import date as dt
from general.productCarriers import dx, parcelforce
from general.functions import formatDateOpposite as formatDate
from general.productNames import barToSku



def upload(orders):
    sheetNum = 1
    workbook = xlsx.load_workbook("dataPOST/D2c Stock and Sales Sheet.xlsx")
    sheets = workbook.sheetnames
    sheetName = sheets[sheetNum]
    sheet = workbook[sheetName]

    lines = []

    for order in orders:
        if order["products"]:
            for prod in order["products"]:
                if prod in dx or prod in parcelforce:
                    info = {"Date":"","SKU":"","Product":"","Order/Reference":"","Movement Type":"Sale / Order","Quantity":""}
                    info["Date"] = formatDate(order["orderDate"])
                    info["Order/Reference"] = order["custPO"]
                    info["Product"] = prod
                    info["Quantity"] = order["products"][prod][0]
                    info["SKU"] = barToSku[order["products"][prod][1]]
                    info["Retailer"] = order["accName"]

                    lines.append(info)

    row = max(cell.row for cell in sheet["A"] if cell.value is not None) + 1

    merged_range = list(sheet.merged_cells.ranges)
    for merged_cell_range in merged_range:
        sheet.unmerge_cells(str(merged_cell_range))

    for line in lines:
        sheet[f"A{row}"] = line["Date"]
        sheet[f"B{row}"] = str(line["SKU"])
        sheet[f"C{row}"] = str(line["Product"])
        sheet[f"D{row}"] = line["Retailer"]
        sheet[f"E{row}"] = str(line["Order/Reference"])
        sheet[f"F{row}"] = line["Movement Type"]
        sheet[f"G{row}"] = int(line["Quantity"])
        row += 1

    workbook.save("dataPOST/D2c Stock and Sales Sheet.xlsx")